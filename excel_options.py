import xlrd
import xlwt
import xlsxwriter
import openpyxl
# filename = "C:/Users/admin/Desktop/修改的表单/1.  桥梁总体外观鉴定检查记录表(外观记录表8.2.2)(SG).xlsx"
#
# data = xlrd.open_workbook(filename)
# table = data.sheets()[0]
#
# nrows = table.nrows
# ncols = table.ncols
# # for i in range(0, nrows):
# #     rowValues = table.row_values(i)#读取某一行数据
# #     for item in rowValues:
# #         print (item)
# for i in range(0, ncols):
#     colValues = table.col_values(i)#读取某一列数据
#     for item in colValues:
#         print(item)

# workbook = xlwt.Workbook()
# sheet = workbook.add_sheet('chenmeng')
# sheet.write(0, 1, 'test text')
# workbook.save('c:/users/admin/desktop/test.xls')
#
# workbook = openpyxl.load_workbook(r"C:\\Users\\admin\Desktop\修改的表单\1.  桥梁总体外观鉴定检查记录表(外观记录表8.2.2)(SG).xlsx")
# sheetnames = workbook.get_sheet_names()#获得表单名字

# print(sheetnames)
#
# sheet = workbook.get_sheet_by_name(sheetnames[0])
#
# print(sheet.cell(row=3, column=3).value)
#
# sheet['A1'] = '47'
#
# workbook.save("c:/users/admin/desktop/桥梁外观记录表01.xlsx")

# def get_chart(series):
#
#     chart = workbook.add_chart({'type': 'line'})
#
#     for ses in series:
#         name = ses["name"]
#         values = ses["values"]
#         chart.add_series({
#             'name': name,
#             'categories': 'chenmeng!A2:A10',
#             'values': values,
#         })
#     chart.set_size({'width':700, 'height':350})
#     return chart
#
# if __name__ == '__main__':
#     workbook = xlsxwriter.Workbook(r"C:\Users\admin\Desktop\1.  桥梁总体外观鉴定检查记录表(外观记录表8.2.2)(SG).xlsx")
#     worksheet = workbook.add_worksheet("chenmeng")
#     bold = workbook.add_format({'bold':1})
#     headings = ['日期', '平均值', '性别']
#     colors = ['red', 'blue', 'green']
#
#     worksheet.write_row('A1', headings)
#     worksheet.write_column('C2', colors)
#     index = 0
#     for row in range (1, 10):
#         for col in [0, 1, 2]:
#             worksheet.write(row, col, index)
#             index += 1
#
#     series = [{"name":"平均值", "values":"chenmeng!B2:B10",}, {'name':'性别', 'values':'chenmeng!C2:C10'}]
#     chart = get_chart(series)
#     # chart.set_x_axis({'name':'日期', 'values':'A2:A10'})
#     chart.set_title({'name':'每日页面分享数据'})
#     # chart.set_legend({'name': 'ceshi'})
#     worksheet.insert_chart('H7', chart)
#     workbook.close()

import win32com.client  #导入脚本模块

ExcelApp = win32com.client.DispatchEx("Excel.Application") #载入Excel模块
ExcelApp.Visible = True #显示excel应用程序

wBook = ExcelApp.Workbooks.Add() #新建空文件，每个文件系统默认建立3个空表
nSheet = wBook.Worksheets.Count #获取文件的表格数， 缺省为3
wSheet = wBook.Worksheets(1) #打开指定工作表，注意序号从1开始
wSheet = wBook.Worksheets.Add() #增加新表，新表为第4个表，相当于wBook.Worksheets(4)

wSheet.Name = "新建表格" #或者wBook.Worksheets(4).Name = "新建表格"

wBook.SaveAs(r"c:\users\admin\desktop\测试表格.xlsx") #另存文件，注意直接保存用下面一行的
# wBook.Save("c:/users/admin/desktop/testchart.xlsx")
wBook.Close()

#打开和关闭xls文件
#先指定路径名称
xlsPathName = r"C:\Users\admin\Desktop\替换筑业\第二批\4 桥梁工程（第四册）（定版）-20180807\1桥梁工程\1施工单位用表\1桥梁评定表（114张）\评定表\评定表8.2.2  桥梁总体评定(SG).xls"
wBook = ExcelApp.Workbooks.Open(xlsPathName) #打开指定文件（账本）
wSheet = wBook.Worksheets(1) #打开指定工作表（注意序号从1开始），或者wSheet.Workbooks("sheet1") 必须要使用准确的工作表名字

#...中间操作...
wBook.Close()

#3、页面设置

wSheet.PageSetup.PaperSize = 9 #设置纸张大小 A3=8 , A4=9(与word不同)
wSheet.PageSetup.Orientation = 1 #设置页面方向，纵向=1， 横向=2（与word不同）

wSheet.PageSetup.TopMargin = 3*28.35  #页边距上 = 3cm， 1cm=28.35pt
wSheet.PageSetup.BottomMargin = 3*28.35
wSheet.PageSetup.LeftMargin = 2.5*28.35
wSheet.PageSetup.RightMargin = 2.5*28.35

wSheet.PageSetup.CenterHorizontally = True  #表格打印位置水平居中
wSheet.PageSetup.CenterVertically = False  #表格打印位置垂直不居中（最后一页不好看）

wSheet.PageSetup.HeaderMargin = 2*28.35  #设置页眉位置=2cm （距离上边界）
wSheet.PageSetup.FooterMargin = 1*28.35  #设置页脚位置 = 1cm（距离下边界）

wSheet.PageSetup.PrintTitleRows = "$1:$2" #设置表格标题行
wSheet.PageSetup.CenterHeader = r"&\'黑体'\&15表格名称"
wSheet.PageSetup.CenterFooter = "第&P页，共&N页"

wSheet.Rows(5).PageBreak = -4135  #在第5行之前插入分页符

#&"黑体"&15   #设置字体、字号、颜色
#&B&I&U   #设置字体加黑、倾斜、下划线
#&L&C&R   #设置左中右对齐

#4，单元格操作

cv = wSheet.Cells(1, 1).Value   #获取单元格数值
wSheet.Cells(1, 1).Interior.Color = 0xff00ff   #设置单元格背景色
cel = wSheet.Cells(2, 2)   #获取单元格对象
cv = cel.Offset(3, 3).Value   #获取偏移后的单元格为（a1+a2-1, b1+b2-1）

#5 行列操作

wSheet.Rows.AutoFit()  #自动适应行
wSheet.Rows(1).Delete()  #删除第1行

wSheet.Columns.AutoFit()  #自动适应列
wSheet.Columns(1).Delete()
wSheet.Columns(1).ColumnWidth = 30  #设置列宽
wSheet.Columns(1).NumberFormatLocal = "000000"  #设置数值格式

wSheet.Rows("2:2").Select()  #必须选择第二行，才能冻结第1行
ExcelApp.ActiveWindow.FreezePanes = True  #冻结第一行


#6 遍历工作表的所有单元格

nRow = wSheet.usedrange.rows.count  #获取指定工作表的行数
nCol = wSheet.usedrange.columns.count  #获取指定工作表的列数

for i in range(1, nRow+1):
    for j in range(1, nCol+1):
      break
    break

#7 搜索指定数值

cel = wSheet.Columns(2).Find(123)   #在第2列查找某整数值123
if cel:
    adr = cel.Address   #获取该单元格的首地址，以便退出循环
    while True:
        cel = wSheet.Columns(2).FindNext(cel)
        if cel.Address == adr:
            break

#格式设置

wSheet.Cells.Font.Name = "Arial"  #设置字体
wSheet.Cells.Font.Size = 10
wSheet.Cells.Font.Bold = True
wSheet.Cells.Font.Italic = True
wSheet.usedrange.Column.AutoFit  #所有列自动调整宽度
wSheet.Cells.HorizontalAlignment = 2   #设置左对齐（1=两端，2=左，3=中，4=右）

ran = wSheet.Range(wSheet.Cells(1, 1), wSheet.Cells(nRow, nCol))

ran.Hyperlinks.Delete()  #删除指定范围的链接
ran.Font.Name = "宋体"  #设置字体
ran.Font.Size =10
ran.Font.Bold = False
ran.Font.Italic = True
ran.HorizontalAlignment = -4108  #水平对齐
ran.VerticalAlignment = -4108   #垂直对齐

#绘制表格线

ran = wSheet.Range("A1:D5")  #设置处理范围
ran.Borders.LineStyle = 1  #设置线型为实线（1=实线，4=点划线，-4142=无线条，-4119=双线，-4115虚线 ）

ran.Borders(11).LineStyle = -4142 #去除范围内中间竖线（1=实线，4=点划线，-4142=无线条，-4119=双线，-4115虚线）

#行列宽度设置

wSheet.Rows(1).RowHeight = 20  #设置第1行的行高
wSheet.Columns(1).ColumnWidth = 10 #设置第1列的列宽

#使用公式进行统计
ran = wSheet.Range("A1:A10")   #设置计算范围
a = ExcelApp.WorksheetFunction.Sum(ran)  #范围数值求和
b = ExcelApp.WorksheetFunction.Max(ran)  #范围最大值

# wSheet.Cells(1,2).Characters(Start:=2, Length:=1).Font.Superscript = True # 设置单元格中第2个字符为上标，如果改为subscript则为下标
# 网上的例子：读取excel文件，并输入字典
#
# def win32Read(filepath, passWords):
#     # win32实现Excel文件读取
#     # Just An Example, Do Not Use It.
#     rst = {}
#     try:
#         pythoncom.CoInitialize() # 如果是多线程操作Microsoft的程序，需要进行初始化

#         xlsApp = win32com.client.DispatchEx('Excel.Application')
#         # 禁用事件
#         xlsApp.EnableEvents = False
#         # 禁止弹窗
#         xlsApp.DisplayAlerts = False
#
#         # 注: 当使用密码时,前边的几个参数都必须存在
#         wb = xlsApp.Workbooks.Open(filepath, UpdateLinks=3, ReadOnly=False, Format=None, Password=passWords)
#
#         # 屏蔽弹窗
#         wb.Checkcompatibility = False
#         # 1:打开宏，2:禁用宏
#         wb.RunAutoMacros(2)
#
#         for sheetObj in wb.Worksheets:
#             datatupe = sheetObj.UsedRange.Value
#             if not datatupe:
#                 continue
#             else:
#                 datatupe = [list(linedata) for linedata in datatupe]
#             rst[sheetObj.name] = datatupe
#         return rst
#     except Exception, e:
#         print
#         unicode(e)
#         # 具体处理看情况
#     finally:
#         try:
#             xlsApp.DisplayAlerts = False
#             wb.Close(SaveChange=False)
#             xlsApp.DisplayAlerts = True
#         except:
#             pass
#         try:
#             xlsApp.Application.Quit()
#             del xlsApp
#         except:
#             pass
#         pythoncom.CoUninitialize()  # 完成操作后，释放资源

